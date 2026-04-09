#!/usr/bin/env python3
"""V7 Padel — Fuzzy matching of Matchpoint bookings to customers.
Reads out(11).xlsx (customers) and out(12).xlsx (bookings),
matches booking text to customer names, reports coverage.

STAFF EXCLUSION: Gulcan Yanar (former trainer) — 253 bookings are work records, not client.
Mark as status=staff_former, exclude from client analytics."""

import openpyxl, re, unicodedata
from collections import Counter, defaultdict

# Staff/former staff to exclude from CLIENT matching (their bookings are work records)
STAFF_EXCLUSIONS = {
    'gulcan yanar',
}

# ---- LOAD ----
wb1 = openpyxl.load_workbook('C:/Users/volod/Downloads/out (11).xlsx', read_only=True)
cust_rows = list(wb1.active.iter_rows(values_only=True))[1:]
wb2 = openpyxl.load_workbook('C:/Users/volod/Downloads/out (12).xlsx', read_only=True)
book_rows = list(wb2.active.iter_rows(values_only=True))[1:]

def norm(s):
    if not s: return ''
    s = str(s).replace('\xa0', ' ').strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[?]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# Build customer index: norm_name -> (code, original)
cust_index = {}
cust_by_first = defaultdict(list)

for r in cust_rows:
    code = r[0]
    name = str(r[2] or '').replace('\xa0', ' ').strip()
    last1 = str(r[3] or '').replace('\xa0', ' ').strip()
    last2 = str(r[4] or '').replace('\xa0', ' ').strip()
    if not name: continue

    parts = [name]
    if last1: parts.append(last1)
    full = ' '.join(parts)
    nfull = norm(full)

    if nfull and len(nfull) > 3:
        # Skip staff exclusions from client index
        if nfull in STAFF_EXCLUSIONS:
            continue
        cust_index[nfull] = (code, full)

    # Also index by reversed order (last first)
    if last1:
        rev = f'{norm(last1)} {norm(name)}'
        if rev != nfull:
            cust_index[rev] = (code, full)

    # First name index for partial match
    nfirst = norm(name)
    if nfirst and len(nfirst) > 2:
        cust_by_first[nfirst].append((code, full, nfull))

    # With both last names
    if last2:
        full2 = full + ' ' + last2
        nfull2 = norm(full2)
        cust_index[nfull2] = (code, full2)

sorted_names = sorted(cust_index.keys(), key=len, reverse=True)

# ---- PATTERNS ----
SPECIAL_TRAINING = re.compile(r'(?:spesial|special)\s+training\s+for\s+(\w+)', re.I)
FREE_TRAINING_NAME = re.compile(r'free trainings?\s+for\s+new\s+clients\.\s*([\w\s]+?)\s*\.', re.I)
VS_PATTERN = re.compile(r'([\w\s]+?)\s+VS\s+([\w\s]+?)$', re.I)

# ---- CLASSIFY ----
total = len(book_rows)
results = []
cat_counts = Counter()
matched_customers = set()

for r in book_rows:
    num = r[0]
    text_raw = str(r[1] or '').replace('\xa0', ' ').strip()
    ntext = norm(text_raw)

    if not ntext:
        cat_counts['empty'] += 1
        results.append((num, None, None, None, 'empty'))
        continue

    # 0. Staff exclusion check — if text IS a staff name, mark as staff
    if ntext in STAFF_EXCLUSIONS:
        cat_counts['staff'] += 1
        results.append((num, None, None, None, 'staff'))
        continue
    # Also check if staff name is substring of text
    staff_found = False
    for sname in STAFF_EXCLUSIONS:
        if sname in ntext:
            cat_counts['staff'] += 1
            results.append((num, None, None, None, 'staff'))
            staff_found = True
            break
    if staff_found:
        continue

    # 1. Exact full match
    if ntext in cust_index:
        code, orig = cust_index[ntext]
        cat_counts['exact'] += 1
        matched_customers.add(code)
        results.append((num, code, orig, 'high', 'exact'))
        continue

    # 2. Substring match (longest first, min 6 chars)
    found_sub = None
    for cn in sorted_names:
        if len(cn) >= 6 and cn in ntext:
            code, orig = cust_index[cn]
            found_sub = (code, orig, 'medium', 'substring')
            break

    if found_sub:
        cat_counts['substring'] += 1
        matched_customers.add(found_sub[0])
        results.append((num, *found_sub))
        continue

    # 3. VS pattern - tournament matchups
    vs_m = VS_PATTERN.search(text_raw)
    if vs_m:
        found = False
        for g in [vs_m.group(1).strip(), vs_m.group(2).strip()]:
            ng = norm(g)
            if ng in cust_index:
                code, orig = cust_index[ng]
                cat_counts['vs_match'] += 1
                matched_customers.add(code)
                results.append((num, code, orig, 'medium', 'vs_match'))
                found = True
                break
            # Try substring within VS group
            for cn in sorted_names:
                if len(cn) >= 6 and cn in ng:
                    code, orig = cust_index[cn]
                    cat_counts['vs_match'] += 1
                    matched_customers.add(code)
                    results.append((num, code, orig, 'low', 'vs_match'))
                    found = True
                    break
            if found:
                break
        if found:
            continue

    # 4. Special training patterns
    sp_m = SPECIAL_TRAINING.search(text_raw)
    if sp_m:
        fname = norm(sp_m.group(1))
        if fname in cust_by_first and len(cust_by_first[fname]) == 1:
            code, orig, _ = cust_by_first[fname][0]
            cat_counts['training_name'] += 1
            matched_customers.add(code)
            results.append((num, code, orig, 'low', 'training_name'))
            continue

    # 5. Free training name
    ft_m = FREE_TRAINING_NAME.search(text_raw)
    if ft_m:
        fname = norm(ft_m.group(1))
        if fname in cust_index:
            code, orig = cust_index[fname]
            cat_counts['free_training'] += 1
            matched_customers.add(code)
            results.append((num, code, orig, 'medium', 'free_training'))
            continue

    # 6. Categorize unmatched by type
    nl = ntext
    if any(kw in nl for kw in ['mexicano','americano','ranked','korzinka','korsinka',
            'korzynka','morning coffee','evening coffee','sunrise','premier',
            'padel&tea','lady','ladies','beginners']):
        cat_counts['tournament_event'] += 1
        results.append((num, None, None, None, 'tournament_event'))
    elif re.search(r'(adult|kids?|individual|group|pair)\s+(group\s+)?training', nl):
        cat_counts['training'] += 1
        results.append((num, None, None, None, 'training'))
    elif any(kw in nl for kw in ['camp bob','amp bob','bobs camp','training with bob',
            'free training']):
        cat_counts['training'] += 1
        results.append((num, None, None, None, 'training'))
    elif nl == 'camp':
        cat_counts['training'] += 1
        results.append((num, None, None, None, 'training'))
    elif re.match(r'match(\s|$)', nl):
        cat_counts['match_generic'] += 1
        results.append((num, None, None, None, 'match_generic'))
    elif re.match(r'booking\s+\d', nl):
        cat_counts['booking_generic'] += 1
        results.append((num, None, None, None, 'booking_generic'))
    elif any(kw in nl for kw in ['event weekend','major cup','tournament','play hard',
            'padel vs tennis','padel&tennis','kids padel cup','training camp']):
        cat_counts['event_block'] += 1
        results.append((num, None, None, None, 'event_block'))
    elif 'maintenance' in nl:
        cat_counts['maintenance'] += 1
        results.append((num, None, None, None, 'maintenance'))
    else:
        cat_counts['unmatched'] += 1
        results.append((num, None, None, None, 'unmatched'))

# ---- REPORT ----
total_matched = sum(1 for r in results if r[1] is not None)
print('=' * 60)
print('FUZZY MATCHING REPORT - V7 Padel Bookings')
print('=' * 60)
print(f'Total bookings:       {total}')
print(f'MATCHED to customer:  {total_matched} ({total_matched/total*100:.1f}%)')
print(f'Unique customers:     {len(matched_customers)} / {len(cust_rows)}')
print()
print('Category breakdown:')
for cat, cnt in cat_counts.most_common():
    pct = cnt / total * 100
    is_match = cat in ('exact','substring','vs_match','training_name','free_training')
    marker = '+' if is_match else ' '
    print(f'  {marker} {cat:20s} {cnt:5d}  ({pct:.1f}%)')

print()
print('Confidence distribution of matched:')
conf = Counter(r[3] for r in results if r[1] is not None)
for c, cnt in conf.most_common():
    print(f'  {c}: {cnt}')

# Show remaining unmatched
um_samples = []
for brow, res in zip(book_rows, results):
    if res[4] == 'unmatched':
        um_samples.append(str(brow[1] or ''))
    if len(um_samples) >= 30:
        break

print(f'\nStill unmatched: {cat_counts.get("unmatched", 0)}')
print('Samples:')
for t in um_samples:
    print(f'  {t}')

# ---- COVERAGE ANALYSIS ----
print('\n' + '=' * 60)
print('COVERAGE ANALYSIS')
print('=' * 60)

def safe_float(v):
    try: return float(str(v).replace(',','').replace('\xa0','').strip() or '0')
    except: return 0.0

# How many bookings have Amount > 0 (paid) vs matched?
paid_total = sum(1 for r in book_rows if safe_float(r[11]) > 0)
paid_matched = 0
for brow, res in zip(book_rows, results):
    if res[1] is not None and safe_float(brow[11]) > 0:
        paid_matched += 1

print(f'Paid bookings (Amount>0):    {paid_total}')
print(f'Paid + matched:              {paid_matched} ({paid_matched/max(paid_total,1)*100:.1f}%)')
print(f'Paid unmatched:              {paid_total - paid_matched}')

# Bookings per matched customer
cust_booking_count = Counter()
for res in results:
    if res[1] is not None:
        cust_booking_count[res[1]] += 1

if cust_booking_count:
    counts = sorted(cust_booking_count.values(), reverse=True)
    print(f'\nBookings per matched customer:')
    print(f'  Max: {counts[0]}, Median: {counts[len(counts)//2]}, Min: {counts[-1]}')
    print(f'  Top 10 most active:')
    for code, cnt in cust_booking_count.most_common(10):
        # find name
        name = next((r[2] for r in results if r[1] == code), '?')
        print(f'    {name}: {cnt} bookings')

# ---- EXPORT CSV ----
import csv, os
out_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'booking_matches.csv')
os.makedirs(os.path.dirname(out_path), exist_ok=True)

with open(out_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['booking_number', 'customer_code', 'customer_name', 'confidence', 'method'])
    for res in results:
        w.writerow(res)

print(f'\nExported to: {os.path.abspath(out_path)}')
print(f'Total rows: {len(results)}')
